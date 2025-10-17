# Importando o openpyxl
import openpyxl

# Nome da planilha de Excel
nome = 'planilha.xlsx'

# Obtendo os dados
def obter_dados_excel(nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)

    return dados

print(obter_dados_excel(nome))

# Deletar produto
def deletar_linha_por_nome(nome_produto, nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if str(row[0]) == nome_produto:
            # Obter o número da linha e deletar a linha inteira
            linha = contador
            sheet.delete_rows(linha)
            break

        # Incrementando o contador
        contador += 1

    # Salvando as mudanças na planilha
    wb.save(nome_planilha)

deletar_linha_por_nome('Carro', nome)