# Importando o openpyxl
import openpyxl

# Definindo as variáveis
preco_compra = float(input("Digite o preço da compra: "))
preco_venda = float(input("Digite o preço da venda: "))
custos_adicionais = float(input("Digite os custos adicionais (opcional): "))
custo_medio_frete = float(input("Digite o custo médio de frete: "))

# Calculando o lucro líquido
lucro_liquido = preco_venda - preco_compra - custos_adicionais - custo_medio_frete

# Imprimindo o resultado
print(f"O lucro líquido é de R${lucro_liquido:.2f}")

# Calculando a margem de lucro
margem_lucro = (lucro_liquido / preco_venda) * 100

# Imprimindo o resultado
print(f"A margem de lucro é de {margem_lucro:.2f}%")

# Salvando os resultados em uma planilha do Excel
nome_produto = 'Produto 1'

resumo = [nome_produto, preco_compra, preco_venda, custos_adicionais, custo_medio_frete, lucro_liquido, margem_lucro]

# Carregando a planilha existente ou criando uma nova planilha se ela não existe
try:
    wb = openpyxl.load_workbook('planilha.xlsx')
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Resultado da Calculadora de Lucro'

    # Adicionando o cabeçalho
    sheet['A1'] = 'Nome do Produto'
    sheet['B1'] = 'Preço de Compra'
    sheet['C1'] = 'Preço de Venda'
    sheet['D1'] = 'Custos Adicionais'
    sheet['E1'] = 'Custo Médio de Frete'
    sheet['F1'] = 'Lucro Líquido'
    sheet['G1'] = 'Margem de Lucro (%)'

    # Adicionando os valores na planilha
    row = sheet.max_row + 1
    sheet[f'A{row}'] = nome_produto
    sheet[f'B{row}'] = preco_compra
    sheet[f'C{row}'] = preco_venda
    sheet[f'D{row}'] = custos_adicionais
    sheet[f'E{row}'] = custo_medio_frete
    sheet[f'F{row}'] = lucro_liquido
    sheet[f'G{row}'] = margem_lucro

    # Salvando a planilha
    wb.save('planilha.xlsx')